from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from .models import Shipment, Message


def export_to_excel(request):
    """
    Web interface uchun Excel export
    """
    days = request.GET.get('days', '1')

    try:
        days = int(days)
    except ValueError:
        days = 1

    end_date = timezone.now()
    start_date = end_date - timedelta(days=days)

    # ✅ TO'G'RILANDI: Shipment va Message modellaridan foydalanish
    shipments = Shipment.objects.filter(
        message__date__gte=start_date,
        message__date__lte=end_date
    ).select_related('message__channel').order_by('-message__date')

    wb = create_excel_workbook(shipments, days)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=telegram_messages_{days}_kun.xlsx'

    wb.save(response)
    return response


def create_excel_workbook(shipments, days):
    """
    Excel workbook yaratish
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Shipments ({days} kun)"

    headers = ['№', 'Kanal', 'Xabar', 'Origin', 'Destination', 'Yuk turi',
               'Transport', 'To\'lov', 'Telefon', 'Sana']
    ws.append(headers)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    for idx, shipment in enumerate(shipments, start=1):
        msg = shipment.message
        ch = msg.channel

        row = [
            idx,
            ch.title if ch else 'Noma\'lum',
            msg.text[:100] + '...' if msg.text and len(msg.text) > 100 else (msg.text or ''),
            shipment.origin or '-',
            shipment.destination or '-',
            shipment.cargo_type or '-',
            shipment.truck_type or '-',
            shipment.payment_type or '-',
            shipment.phone or '-',
            msg.date.strftime('%Y-%m-%d %H:%M') if msg.date else '-'
        ]
        ws.append(row)

        for cell in ws[idx + 1]:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    column_widths = [5, 25, 50, 20, 20, 20, 20, 20, 20, 20]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    ws.row_dimensions[1].height = 25
    for row in range(2, len(shipments) + 2):
        ws.row_dimensions[row].height = 20

    return wb


def generate_excel_file(shipments, days):
    """
    Telegram bot uchun BytesIO qaytaradi
    """
    wb = create_excel_workbook(shipments, days)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_shipments_workbook_bytes(days=1):
    """
    bot_service.py uchun funksiya - async kontekstdan chaqiriladi
    """
    from asgiref.sync import sync_to_async

    end_date = timezone.now()
    start_date = end_date - timedelta(days=days)

    # ✅ TO'G'RILANDI: Shipment modelidan foydalanish
    shipments = list(Shipment.objects.filter(
        message__date__gte=start_date,
        message__date__lte=end_date
    ).select_related('message__channel').order_by('-message__date'))

    return generate_excel_file(shipments, days)


def export_to_json(request):
    """
    JSON formatda export
    """
    import json

    days = int(request.GET.get('days', '1'))
    end_date = timezone.now()
    start_date = end_date - timedelta(days=days)

    # ✅ TO'G'RILANDI: Shipment modelidan foydalanish
    shipments = Shipment.objects.filter(
        message__date__gte=start_date,
        message__date__lte=end_date
    ).select_related('message__channel').order_by('-message__date')

    data = []
    for shipment in shipments:
        msg = shipment.message
        ch = msg.channel
        data.append({
            'channel': ch.title if ch else None,
            'text': msg.text,
            'origin': shipment.origin,
            'destination': shipment.destination,
            'cargo_type': shipment.cargo_type,
            'truck_type': shipment.truck_type,
            'payment_type': shipment.payment_type,
            'phone': shipment.phone,
            'date': msg.date.isoformat() if msg.date else None
        })

    response = HttpResponse(
        json.dumps(data, ensure_ascii=False, indent=2),
        content_type='application/json'
    )
    response['Content-Disposition'] = f'attachment; filename=telegram_messages_{days}_kun.json'

    return response