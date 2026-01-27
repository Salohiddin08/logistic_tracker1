import asyncio
import re
import threading
from django.conf import settings
from django.contrib.auth import logout
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST
from asgiref.sync import async_to_sync

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import TelegramSession, Channel, Message, Shipment
from .telethon_client import get_client, get_channels, get_messages
from .utils import save_messages_json, parse_shipment_text
from .bot_service import send_export_now
from telethon import TelegramClient
from telethon.sessions import StringSession
from telethon.errors import SessionPasswordNeededError


def _run_async_in_thread(coro):
    """
    Helper function: async coroutine ni alohida threadda ishga tushirish
    """
    result = {'data': None, 'error': None}
    
    def run_in_thread():
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            result['data'] = loop.run_until_complete(coro)
        except Exception as e:
            result['error'] = e
        finally:
            loop.close()
    
    thread = threading.Thread(target=run_in_thread)
    thread.start()
    thread.join()
    
    if result['error']:
        raise result['error']
    return result['data']


# ==================== HELPER FUNCTIONS ====================

def highlight_text(text, keywords):
    """
    Text ichidagi keywordslarni sariq rangda highlight qilish
    """
    if not text or not keywords:
        return text
    
    # Har bir keyword uchun case-insensitive replace
    highlighted = text
    for keyword in keywords:
        if keyword.strip():
            # Regex bilan case-insensitive replace
            pattern = re.compile(re.escape(keyword), re.IGNORECASE)
            highlighted = pattern.sub(
                lambda m: f'<mark style="background-color: yellow;">{m.group()}</mark>',
                highlighted
            )
    
    return highlighted


def _get_tg_credentials():
    api_id = getattr(settings, 'TG_API_ID', None)
    api_hash = getattr(settings, 'TG_API_HASH', None)

    if not api_id or not api_hash:
        raise ValueError("TG_API_ID / TG_API_HASH .env faylda to'ldirilmagan")

    try:
        api_id = int(api_id)
    except Exception as exc:
        raise ValueError("TG_API_ID raqam (int) bo'lishi kerak") from exc

    return api_id, api_hash


# ==================== AUTHENTICATION VIEWS ====================

async def _start_phone_login(phone: str):
    api_id, api_hash = _get_tg_credentials()
    client = TelegramClient(StringSession(), api_id, api_hash)
    await client.connect()
    sent = await client.send_code_request(phone)
    temp_session = client.session.save()
    return temp_session, sent.phone_code_hash


async def _complete_phone_login(temp_session: str, phone: str, code: str, password: str | None, phone_code_hash: str | None):
    api_id, api_hash = _get_tg_credentials()
    client = TelegramClient(StringSession(temp_session), api_id, api_hash)
    await client.connect()

    try:
        await client.sign_in(phone=phone, code=code, phone_code_hash=phone_code_hash)
    except SessionPasswordNeededError:
        if not password:
            raise ValueError("Telegram akkauntingizda 2-bosqichli parol (Two‑step verification) yoqilgan. Parolni kiriting.")
        await client.sign_in(password=password)

    return client.session.save()


def telegram_phone_login(request):
    error = None

    if request.method == 'POST':
        phone = request.POST.get('phone')
        if phone:
            try:
                # Thread ichida async funksiyani ishga tushirish
                temp_session, phone_code_hash = _run_async_in_thread(_start_phone_login(phone))
                    
                request.session['tg_phone'] = phone
                request.session['tg_temp_session'] = temp_session
                request.session['tg_phone_code_hash'] = phone_code_hash
                return redirect('telegram_phone_code')
            except Exception as exc:
                error = str(exc)

    return render(request, 'telegram_login_phone.html', {'error': error})


def telegram_phone_code(request):
    phone = request.session.get('tg_phone')
    temp_session = request.session.get('tg_temp_session')
    phone_code_hash = request.session.get('tg_phone_code_hash')
    if not phone or not temp_session:
        return redirect('telegram_phone_login')

    error = None
    if request.method == 'POST':
        code = request.POST.get('code')
        password = request.POST.get('password') or None
        if code:
            try:
                # Thread ichida async funksiyani ishga tushirish
                string_session = _run_async_in_thread(
                    _complete_phone_login(temp_session, phone, code, password, phone_code_hash)
                )

                api_id = getattr(settings, 'TG_API_ID', '')
                api_hash = getattr(settings, 'TG_API_HASH', '')
                TelegramSession.objects.create(
                    api_id=api_id,
                    api_hash=api_hash,
                    string_session=string_session,
                )
                request.session.pop('tg_phone', None)
                request.session.pop('tg_temp_session', None)
                request.session.pop('tg_phone_code_hash', None)
                return redirect('channels')
            except Exception as exc:
                error = str(exc)

    context = {
        'phone': phone,
        'error': error,
    }
    return render(request, 'telegram_login_code.html', context)


def add_session(request):
    if request.method == 'POST':
        api_id = request.POST.get('api_id')
        api_hash = request.POST.get('api_hash')
        string_session = request.POST.get('string_session')

        TelegramSession.objects.create(
            api_id=api_id,
            api_hash=api_hash,
            string_session=string_session
        )
        return redirect('channels')

    default_api_id = getattr(settings, 'TG_API_ID', '')
    default_api_hash = getattr(settings, 'TG_API_HASH', '')

    context = {
        'default_api_id': default_api_id,
        'default_api_hash': default_api_hash,
    }
    return render(request, 'add_session.html', context)


# ==================== DASHBOARD ====================

def home_view(request):
    if TelegramSession.objects.last():
        return redirect('dashboard')
    return redirect('telegram_phone_login')


def dashboard_view(request):
    today = timezone.localdate()

    shipments = Shipment.objects.select_related('message__channel').filter(message__date__date=today)

    total_today = shipments.count()

    top_origins = (
        shipments
        .values('origin')
        .annotate(total=Count('id'))
        .order_by('-total')[:10]
    )
    top_destinations = (
        shipments
        .values('destination')
        .annotate(total=Count('id'))
        .order_by('-total')[:10]
    )
    top_payments = (
        shipments
        .values('payment_type')
        .annotate(total=Count('id'))
        .order_by('-total')[:10]
    )
    top_cargo = (
        shipments
        .values('cargo_type')
        .annotate(total=Count('id'))
        .order_by('-total')[:10]
    )

    sent = request.GET.get('sent')
    error = request.GET.get('err')

    context = {
        'today': today,
        'total_today': total_today,
        'top_origins': top_origins,
        'top_destinations': top_destinations,
        'top_payments': top_payments,
        'top_cargo': top_cargo,
        'sent': sent,
        'error': error,
    }
    return render(request, 'dashboard.html', context)


@require_POST
def bot_export_view(request):
    try:
        days = int(request.POST.get('days') or 1)
    except Exception:
        days = 1

    try:
        async_to_sync(send_export_now)(days=days)
        return redirect(f"/dashboard/?sent=1")
    except Exception as exc:
        return redirect(f"/dashboard/?err={str(exc)}")


# ==================== 2️⃣ CHANNELS MANAGEMENT (SUBSCRIPTION) ====================
def channels_view(request):
    """
    Barcha kanallarni ko'rsatish + is_tracked bilan boshqarish
    """
    session = TelegramSession.objects.last()
    if not session:
        return redirect('telegram_phone_login')

    async def run():
        client = await get_client(session.api_id, session.api_hash, session.string_session)
        tg_channels = await get_channels(client)
        return tg_channels  # Faqat ma'lumotni qaytarish

    try:
        # Thread ichida async funksiyani ishga tushirish
        channels = _run_async_in_thread(run())
        
        # ✅ Django ORM operatsiyalarini sync kontekstda bajarish
        for ch in channels:
            Channel.objects.get_or_create(
                channel_id=ch['id'],
                defaults={'title': ch['title']}
            )
        
    except Exception as exc:
        return render(
            request,
            'error.html',
            {
                'title': 'Telegram ulanish xatosi',
                'message': "Kanallarni olishda xatolik.",
                'detail': str(exc),
            },
            status=500,
        )

    # DB'dan is_tracked statusini olish
    db_channels = {ch.channel_id: ch for ch in Channel.objects.all()}
    
    for ch in channels:
        ch['is_tracked'] = db_channels.get(ch['id'], Channel()).is_tracked

    return render(request, 'channels.html', {'channels': channels})


@require_POST
def toggle_channel_tracking(request, channel_id):
    """
    Kanalni kuzatishni yoqish/o'chirish
    """
    channel = get_object_or_404(Channel, channel_id=channel_id)
    channel.is_tracked = not channel.is_tracked
    channel.save()
    
    return redirect('channels')


# ==================== FETCH MESSAGES ====================
def fetch_messages_view(request, channel_id):
    session = TelegramSession.objects.last()
    if not session:
        return redirect('add_session')

    async def run():
        client = await get_client(session.api_id, session.api_hash, session.string_session)
        return await get_messages(client, channel_id, limit=100)

    # Thread ichida async funksiyani ishga tushirish
    messages = _run_async_in_thread(run())

    channel_obj, _ = Channel.objects.get_or_create(channel_id=channel_id)

    for m in messages:
        # 1. Asosiy xabarni saqlash yoki olish
        msg_obj, _ = Message.objects.get_or_create(
            channel=channel_obj,
            message_id=m.id,
            defaults={
                'sender_id': getattr(m.from_id, 'user_id', None),
                'sender_name': getattr(m.sender, 'username', None) if m.sender else None,
                'text': m.message,
                'date': m.date,
            },
        )

        # 2. Xabarni tahlil qilish (Endi bu LIST qaytaradi)
        parsed_shipments = parse_shipment_text(m.message or "")

        # 3. Har bir topilgan yukni alohida Shipment sifatida saqlash
# views.py -> fetch_messages_view ichida
        for parsed in parsed_shipments:
            Shipment.objects.update_or_create(
                message=msg_obj,
                origin=parsed.get('origin'),
                destination=parsed.get('destination'),
                phone=parsed.get('phone'), # <--- Filtrga qo'shildi
                defaults={
                    'cargo_type': parsed.get('cargo_type'),
                    'truck_type': parsed.get('truck_type'),
                    'payment_type': parsed.get('payment_type'),
                },
            )
            return redirect('channel_stats', channel_id=channel_id)

# ==================== 1️⃣ MESSAGES WITH TAG SEARCH & HIGHLIGHT ====================

def saved_messages_view(request):
    """
    1️⃣ TAG filter va highlight bilan messages
    """
    messages = Message.objects.select_related('channel').order_by('-date')

    # Sana filtri
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    parsed_from = parse_date(date_from) if date_from else None
    parsed_to = parse_date(date_to) if date_to else None

    if parsed_from:
        messages = messages.filter(date__date__gte=parsed_from)
    if parsed_to:
        messages = messages.filter(date__date__lte=parsed_to)

    # 1️⃣ TAG SEARCH (so'zlar bo'yicha OR logic)
    search_query = request.GET.get('search', '').strip()
    keywords = []
    
    if search_query:
        # Space bilan ajratilgan so'zlar
        keywords = [kw.strip() for kw in search_query.split() if kw.strip()]
        
        # OR logic: kamida bitta so'z bo'lsa ham chiqsin
        q_objects = Q()
        for keyword in keywords:
            q_objects |= Q(text__icontains=keyword)
        
        messages = messages.filter(q_objects)

    # Highlight uchun keywordslarni context'ga yuborish
    paginator = Paginator(messages, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Har bir message uchun highlighted text yaratish
    for msg in page_obj.object_list:
        msg.highlighted_text = highlight_text(msg.text, keywords)

    context = {
        'messages': page_obj.object_list,
        'page_obj': page_obj,
        'date_from': date_from,
        'date_to': date_to,
        'search_query': search_query,
        'keywords': keywords,
    }
    return render(request, 'messages.html', context)


# ==================== 3️⃣ MESSAGE DETAIL VIEW ====================

def message_detail_view(request, message_id):
    """
    3️⃣ Message ustiga bosganda to'liq ko'rinish
    """
    message = get_object_or_404(Message.objects.select_related('channel'), pk=message_id)
    
    try:
        shipment = message.shipment
    except Shipment.DoesNotExist:
        shipment = None
    
    context = {
        'message': message,
        'shipment': shipment,
    }
    return render(request, 'message_detail.html', context)


# ==================== 4️⃣ DUPLICATE DETECTION ====================

def route_duplicates_view(request, channel_id):
    """
    4️⃣ Bir xil yo'nalish bo'yicha dublikatlarni ko'rsatish
    """
    origin = request.GET.get('origin')
    destination = request.GET.get('destination')
    
    shipments = Shipment.objects.filter(
        message__channel__channel_id=channel_id,
        origin=origin,
        destination=destination
    ).select_related('message').order_by('-message__date')
    
    # Dublikat detection (text similarity)
    total_count = shipments.count()
    unique_texts = set()
    duplicate_count = 0
    
    shipments_with_status = []
    for shipment in shipments:
        text = (shipment.message.text or "").strip().lower()
        
        if text in unique_texts:
            is_duplicate = True
            duplicate_count += 1
        else:
            is_duplicate = False
            unique_texts.add(text)
        
        shipments_with_status.append({
            'shipment': shipment,
            'is_duplicate': is_duplicate
        })
    
    unique_count = total_count - duplicate_count
    
    context = {
        'channel_id': channel_id,
        'origin': origin,
        'destination': destination,
        'total_count': total_count,
        'duplicate_count': duplicate_count,
        'unique_count': unique_count,
        'shipments': shipments_with_status,
    }
    return render(request, 'route_duplicates.html', context)


# ==================== EXISTING VIEWS (Updated) ====================

def _get_filtered_shipments(request, channel_id):
    shipments = Shipment.objects.filter(message__channel__channel_id=channel_id)

    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    parsed_from = parse_date(date_from) if date_from else None
    parsed_to = parse_date(date_to) if date_to else None

    if parsed_from:
        shipments = shipments.filter(message__date__date__gte=parsed_from)
    if parsed_to:
        shipments = shipments.filter(message__date__date__lte=parsed_to)

    search_query = request.GET.get('search', '').strip()
    if search_query:
        shipments = shipments.filter(
            Q(origin__icontains=search_query) |
            Q(destination__icontains=search_query) |
            Q(cargo_type__icontains=search_query) |
            Q(truck_type__icontains=search_query) |
            Q(payment_type__icontains=search_query) |
            Q(phone__icontains=search_query)
        )

    return shipments, date_from, date_to


def channel_stats_view(request, channel_id):
    """
    4️⃣ Yo'nalishlar bilan dublikat hisobi
    """
    shipments, date_from, date_to = _get_filtered_shipments(request, channel_id)

    # A → B yo'nalishlar + dublikat hisobi
    route_qs = (
        shipments
        .values('origin', 'destination')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    route_paginator = Paginator(route_qs, 20)
    route_page_number = request.GET.get('route_page', 1)
    route_page_obj = route_paginator.get_page(route_page_number)
    route_stats = route_page_obj.object_list

    cargo_qs = (
        shipments
        .values('cargo_type')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    cargo_paginator = Paginator(cargo_qs, 20)
    cargo_page_number = request.GET.get('cargo_page', 1)
    cargo_page_obj = cargo_paginator.get_page(cargo_page_number)
    cargo_stats = cargo_page_obj.object_list

    truck_qs = (
        shipments
        .values('truck_type')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    truck_paginator = Paginator(truck_qs, 20)
    truck_page_number = request.GET.get('truck_page', 1)
    truck_page_obj = truck_paginator.get_page(truck_page_number)
    truck_stats = truck_page_obj.object_list

    payment_qs = (
        shipments
        .values('payment_type')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    payment_paginator = Paginator(payment_qs, 20)
    payment_page_number = request.GET.get('payment_page', 1)
    payment_page_obj = payment_paginator.get_page(payment_page_number)
    payment_stats = payment_page_obj.object_list

    context = {
        'channel_id': channel_id,
        'total_shipments': shipments.count(),
        'route_stats': route_stats,
        'route_page_obj': route_page_obj,
        'cargo_stats': cargo_stats,
        'truck_stats': truck_stats,
        'payment_stats': payment_stats,
        'cargo_page_obj': cargo_page_obj,
        'truck_page_obj': truck_page_obj,
        'payment_page_obj': payment_page_obj,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'stats.html', context)


# ==================== REMAINING VIEWS ====================

def channel_stats_excel(request, channel_id):
    shipments, date_from, date_to = _get_filtered_shipments(request, channel_id)

    filename_parts = [f"channel_{channel_id}"]
    if date_from:
        filename_parts.append(f"from_{date_from}")
    if date_to:
        filename_parts.append(f"to_{date_to}")
    filename_base = "_".join(filename_parts)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Shipments"

    headers = [
        'channel_id', 'channel_title', 'message_id', 'date',
        'origin', 'destination', 'cargo_type', 'truck_type',
        'payment_type', 'phone',
    ]
    ws.append(headers)

    for shipment in shipments.select_related('message__channel'):
        msg = shipment.message
        ch = msg.channel
        ws.append([
            ch.channel_id,
            ch.title or "",
            msg.message_id,
            msg.date.isoformat() if msg.date else "",
            shipment.origin or "",
            shipment.destination or "",
            shipment.cargo_type or "",
            shipment.truck_type or "",
            shipment.payment_type or "",
            shipment.phone or "",
        ])

    for idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = 18

    wb.save(response)
    return response


def channel_phones_view(request, channel_id):
    shipments, date_from, date_to = _get_filtered_shipments(request, channel_id)

    search_query = request.GET.get('search', '').strip()

    raw_stats = (
        shipments
        .exclude(phone__isnull=True)
        .exclude(phone__exact="")
        .values('phone')
        .annotate(total=Count('id'))
        .order_by('-total')
    )

    if search_query:
        raw_stats = raw_stats.filter(phone__icontains=search_query)

    phone_stats = []
    id_stats = []

    for item in raw_stats:
        phone = item['phone'] or ""
        digits_only = ''.join(ch for ch in phone if ch.isdigit())
        if phone.startswith('+') or len(digits_only) >= 9:
            phone_stats.append(item)
        else:
            id_stats.append(item)

    context = {
        'channel_id': channel_id,
        'phone_stats': phone_stats,
        'id_stats': id_stats,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'phones.html', context)


def channel_phones_excel(request, channel_id):
    shipments, date_from, date_to = _get_filtered_shipments(request, channel_id)

    search_query = request.GET.get('search', '').strip()

    phone_stats = (
        shipments
        .exclude(phone__isnull=True)
        .exclude(phone__exact="")
        .values('phone')
        .annotate(total=Count('id'))
        .order_by('-total')
    )

    if search_query:
        phone_stats = phone_stats.filter(phone__icontains=search_query)

    filename_parts = [f"channel_{channel_id}_phones"]
    if date_from:
        filename_parts.append(f"from_{date_from}")
    if date_to:
        filename_parts.append(f"to_{date_to}")
    if search_query:
        filename_parts.append(f"search_{search_query[:20]}")
    filename_base = "_".join(filename_parts)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Phones"

    headers = ['phone', 'total_shipments']
    ws.append(headers)

    for item in phone_stats:
        ws.append([
            item['phone'],
            item['total'],
        ])

    for idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = 20

    wb.save(response)
    return response


def channel_phone_messages_view(request, channel_id):
    shipments, date_from, date_to = _get_filtered_shipments(request, channel_id)

    phone = request.GET.get('phone') or None
    if phone:
        shipments = shipments.filter(phone=phone)

    shipments = shipments.select_related('message__channel').order_by('-message__date')

    context = {
        'channel_id': channel_id,
        'phone': phone,
        'origin': None,
        'destination': None,
        'cargo_type': None,
        'truck_type': None,
        'payment_type': None,
        'date_from': date_from,
        'date_to': date_to,
        'shipments': shipments,
    }
    return render(request, 'phone_messages.html', context)


def channel_route_messages_view(request, channel_id):
    shipments, date_from, date_to = _get_filtered_shipments(request, channel_id)

    origin = request.GET.get('origin') or None
    destination = request.GET.get('destination') or None

    if origin:
        shipments = shipments.filter(origin=origin)
    if destination:
        shipments = shipments.filter(destination=destination)

    shipments = shipments.select_related('message__channel').order_by('-message__date')

    context = {
        'channel_id': channel_id,
        'origin': origin,
        'destination': destination,
        'cargo_type': None,
        'date_from': date_from,
        'date_to': date_to,
        'shipments': shipments,
    }
    return render(request, 'route_messages.html', context)


def channel_cargo_messages_view(request, channel_id):
    shipments, date_from, date_to = _get_filtered_shipments(request, channel_id)

    cargo_type = request.GET.get('cargo_type') or None
    if cargo_type:
        shipments = shipments.filter(cargo_type=cargo_type)

    shipments = shipments.select_related('message__channel').order_by('-message__date')

    context = {
        'channel_id': channel_id,
        'origin': None,
        'destination': None,
        'cargo_type': cargo_type,
        'date_from': date_from,
        'date_to': date_to,
        'shipments': shipments,
    }
    return render(request, 'route_messages.html', context)


def channel_truck_messages_view(request, channel_id):
    shipments, date_from, date_to = _get_filtered_shipments(request, channel_id)

    truck_type = request.GET.get('truck_type') or None
    if truck_type:
        shipments = shipments.filter(truck_type=truck_type)

    shipments = shipments.select_related('message__channel').order_by('-message__date')

    context = {
        'channel_id': channel_id,
        'origin': None,
        'destination': None,
        'cargo_type': None,
        'truck_type': truck_type,
        'payment_type': None,
        'date_from': date_from,
        'date_to': date_to,
        'shipments': shipments,
    }
    return render(request, 'route_messages.html', context)


def channel_payment_messages_view(request, channel_id):
    shipments, date_from, date_to = _get_filtered_shipments(request, channel_id)

    payment_type = request.GET.get('payment_type') or None
    if payment_type:
        shipments = shipments.filter(payment_type=payment_type)

    shipments = shipments.select_related('message__channel').order_by('-message__date')

    context = {
        'channel_id': channel_id,
        'origin': None,
        'destination': None,
        'cargo_type': None,
        'truck_type': None,
        'payment_type': payment_type,
        'date_from': date_from,
        'date_to': date_to,
        'shipments': shipments,
    }
    return render(request, 'route_messages.html', context)


def export_json(request):
    save_messages_json()
    return HttpResponse("JSON file created successfully!")


def logout_view(request):
    """Oddiy GET orqali ham chiqishni qo'llab-quvvatlaydigan logout."""
    logout(request)
    return redirect('telegram_phone_login')


def channel_search(request, channel_id):
    """Kanal ichida barcha shipmentlar bo'yicha umumiy search."""
    channel = Channel.objects.get(pk=channel_id)
    query = request.GET.get('q', '').strip()

    shipments = Shipment.objects.filter(message__channel=channel)
    if query:
        shipments = shipments.filter(
            Q(origin__icontains=query) |
            Q(destination__icontains=query) |
            Q(cargo_type__icontains=query) |
            Q(truck_type__icontains=query) |
            Q(payment_type__icontains=query) |
            Q(phone__icontains=query)
        )

    return render(request, 'channel_search.html', {
        'channel': channel,
        'shipments': shipments,
        'query': query,
    })


def excel_export_page(request):
    """Excel export sahifasi"""
    return render(request, 'telegram_app/excel_export.html')